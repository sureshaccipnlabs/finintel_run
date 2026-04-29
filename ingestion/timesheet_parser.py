from openpyxl import load_workbook
from datetime import datetime, date
import os
import re

from .normalizer import normalize_record

try:
    from .ai_mapper import ai_map_columns, ai_analyze_sheet, is_ollama_available
    _AI_AVAILABLE = True
except ImportError:
    _AI_AVAILABLE = False
    def ai_map_columns(*a, **kw): return {}
    def ai_analyze_sheet(*a, **kw): return None
    def is_ollama_available(): return False

HOURS_PER_DAY = 8.0

LEAVE_THRESHOLD = 3
MARGIN_LOW_THRESHOLD = 10
UTILISATION_LOW_THRESHOLD = 80

# ── Generic column-name patterns (order matters: most specific first) ────
FIELD_PATTERNS = {
    "name":          ["employee name", "emp name", "resource name", "staff name",
                      "full name", "associate name", "member name", "consultant name",
                      "worker name", "contractor name",
                      "name", "employee", "person", "consultant", "resource",
                      "worker", "contractor", "associate", "staff", "member"],
    "project":       ["project name", "client name", "account name",
                      "project", "client", "account", "engagement", "program",
                      "sow", "assignment"],
    "billing_rate":  ["billing_rate", "billing rate", "bill rate", "rate per hour",
                      "charge rate", "sell rate", "client rate",
                      "rate", "billing"],
    "cost_rate":     ["cost_rate", "cost rate", "ctc rate", "internal rate",
                      "pay rate", "cost", "internal cost", "resource cost",
                      "cost per hour", "cost/hr", "cosst"],
    "actual_hours":  ["actual hours", "total actual", "hours worked", "work hours",
                      "logged hours", "total hours", "worked hours",
                      "actual"],
    "billable_hours":["final billable hours", "final billable", "total billable",
                      "net billable", "billed hours"],
    "max_hours":     ["max hours", "max billable", "expected hours", "budgeted hours",
                      "target hours", "planned hours", "allocated hours",
                      "approved", "budgeted"],
    "leaves":        ["leave days", "leaves", "vacation days", "vacation",
                      "time off", "pto", "absent", "days off"],
    "working_days":  ["working days", "work days", "days worked",
                      "business days", "workdays"],
}

# Prevent false positives: if a cell matches field X but contains one of
# these keywords, skip it — it likely belongs to another field.
FIELD_EXCLUSIONS = {
    "name":          ["project", "client", "account", "engagement", "program",
                      "sow", "role"],
    "project":       ["rate", "hour", "cost", "amount", "days", "billing",
                      "role", "sow", "designation", "title"],
    "billing_rate":  ["final", "amount", "total", "cost", "ctc"],
    "cost_rate":     ["billing", "bill", "charge", "sell"],
    "actual_hours":  ["max", "approved", "budgeted", "billable"],
    "billable_hours":["max", "approved", "budgeted"],
    "max_hours":     [],
}

SKIP_NAMES = {"name", "total", "sl no.", "sl no", "grand total", "totals", ""}


# ── Primitives ────────────────────────────────────────────────────────────
def clean(v):
    return str(v).strip() if v else ""


def is_date(v):
    return isinstance(v, (datetime, date))


def _find_col(header, label):
    for j, v in enumerate(header):
        if v and label in clean(v).lower():
            return j
    return None


def _smart_find_col(header, field):
    """Try every pattern for *field* until one matches a column.
    Skips columns whose text contains an exclusion keyword for this field."""
    patterns = FIELD_PATTERNS.get(field, [field])
    exclusions = FIELD_EXCLUSIONS.get(field, [])
    for pattern in patterns:
        for j, v in enumerate(header):
            if not v:
                continue
            cell = clean(v).lower()
            if pattern in cell:
                if exclusions and any(ex in cell for ex in exclusions):
                    continue
                return j
    return None


def _safe_float(row, idx):
    if idx is None or idx >= len(row) or row[idx] is None:
        return 0
    try:
        return float(row[idx])
    except (ValueError, TypeError):
        return 0


_LEAVE_MARKERS = {"off", "l", "leave", ""}
_HOLIDAY_MARKERS = {"ph", "holiday", "public holiday"}
_WEEKDAY_MARKERS = {
    "mon", "monday", "tue", "tues", "tuesday", "wed", "wednesday",
    "thu", "thur", "thurs", "thursday", "fri", "friday", "sat", "saturday",
    "sun", "sunday"
}


def _is_daily_date_col(v):
    if is_date(v):
        return True
    if isinstance(v, (int, float)):
        iv = int(v)
        return float(iv) == float(v) and 1 <= iv <= 31

    s = clean(v).strip().lower()
    if not s:
        return False
    if s in _WEEKDAY_MARKERS:
        return True
    if re.match(r"^\d{1,2}$", s):
        return 1 <= int(s) <= 31
    if re.match(r"^\d{1,2}[/-]\d{1,2}([/-]\d{2,4})?$", s):
        return True
    if re.match(r"^\d{4}[/-]\d{1,2}[/-]\d{1,2}$", s):
        return True
    return False


def _get_daily_date_columns(header):
    cols = [i for i, v in enumerate(header or []) if _is_daily_date_col(v)]
    return cols if len(cols) >= 5 else []


def _compute_daily_metrics(row, day_cols):
    working_days = 0
    leave_days = 0
    holiday_days = 0
    actual_hours = 0.0

    for idx in day_cols:
        if idx >= len(row):
            continue

        val = row[idx]
        if val is None:
            leave_days += 1
            continue

        if isinstance(val, (int, float)):
            hrs = float(val)
            if hrs > 0:
                actual_hours += hrs
                working_days += 1
            continue

        s = clean(val).strip()
        low = s.lower()

        if low in _LEAVE_MARKERS:
            leave_days += 1
            continue
        if low in _HOLIDAY_MARKERS:
            holiday_days += 1
            continue

        try:
            hrs = float(low)
            if hrs > 0:
                actual_hours += hrs
                working_days += 1
        except (ValueError, TypeError):
            continue

    return {
        "actual_hours": round(actual_hours, 2),
        "working_days": working_days,
        "leave_days": leave_days,
        "holiday_days": holiday_days,
    }


def _safe_int(row, idx):
    if idx is None or idx >= len(row) or row[idx] is None:
        return 0
    try:
        return int(float(row[idx]))
    except (ValueError, TypeError):
        return 0


def _has_cell_value(row, idx):
    if idx is None or idx >= len(row):
        return False
    val = row[idx]
    if val is None:
        return False
    if isinstance(val, str):
        return val.strip() != ""
    return True


# ── Column mapping (patterns + AI) ──────────────────────────────────────
_ai_col_cache = {}


def _map_columns(header, fields=None):
    """
    Build field→col_index mapping for a header row.
    Step 1: pattern matching.  Step 2: AI fills remaining gaps.
    Returns dict {field: col_index_or_None}.
    """
    if fields is None:
        fields = list(FIELD_PATTERNS.keys())
    mapped = {f: _smart_find_col(header, f) for f in fields}

    # Resolve conflicts: if two fields map to the same column, keep the more
    # specific one (max_hours beats billable_hours on the same column)
    idx_to_field = {}
    for f in ["name", "project", "max_hours", "actual_hours", "billing_rate",
              "cost_rate", "billable_hours", "leaves", "working_days"]:
        idx = mapped.get(f)
        if idx is not None:
            if idx in idx_to_field:
                mapped[f] = None          # later/less-specific field loses
            else:
                idx_to_field[idx] = f

    # AI fallback only when essential fields (within requested set) are missing
    essential = {"name", "actual_hours"}
    requested_essential = essential.intersection(set(fields))
    missing_essential = [f for f in requested_essential if mapped.get(f) is None]
    if missing_essential and _AI_AVAILABLE and is_ollama_available():
        critical = [f for f in fields if mapped.get(f) is None]
        if critical:
            key = tuple(str(v) for v in header)
            if key not in _ai_col_cache:
                already = {f: i for f, i in mapped.items() if i is not None}
                _ai_col_cache[key] = ai_map_columns(header, already_mapped=already)
            claimed_indices = {i for i in mapped.values() if i is not None}
            for f, idx in _ai_col_cache[key].items():
                if mapped.get(f) is None and idx not in claimed_indices:
                    mapped[f] = idx
                    claimed_indices.add(idx)

    return mapped


# ── Header / section detection ───────────────────────────────────────────
def _is_data_header(row):
    """True when a row looks like a data header (has 'name' + ≥1 other field)."""
    if not row:
        return False
    vals = [clean(v).lower() for v in row]
    has_name = any(
        len(v) < 20 and any(p == v or p in v for p in FIELD_PATTERNS["name"])
        for v in vals if v
    )
    if not has_name:
        return False
    other = 0
    for field, patterns in FIELD_PATTERNS.items():
        if field == "name":
            continue
        if any(v and len(v) < 30 and any(p in v for p in patterns) for v in vals):
            other += 1
    return other >= 1


def _classify_section(header):
    """Return 'fortnight', 'summary', or 'unknown'."""
    has_name  = _smart_find_col(header, "name") is not None
    has_dates = any(is_date(v) for v in header)
    has_rates = (
        _smart_find_col(header, "billing_rate") is not None
        or _smart_find_col(header, "cost_rate") is not None
    )
    has_hours = (
        _smart_find_col(header, "actual_hours") is not None
        or _smart_find_col(header, "max_hours") is not None
    )
    if has_name and has_dates:
        return "fortnight"
    if has_name and (has_rates or has_hours):
        return "summary"
    return "unknown"


# ── Metadata helpers ─────────────────────────────────────────────────────
def _extract_project_from_metadata(rows):
    """Scan the first few rows for a project name label + value pair."""
    for i, row in enumerate(rows[:6]):
        for j, v in enumerate(row):
            label = clean(v).lower()
            if label in ("project", "project name", "client", "client name", "account"):
                # Same row, next non-empty cell
                for k in range(j + 1, min(j + 4, len(row))):
                    cand = clean(row[k])
                    if cand and cand.lower() not in ("project", "project name", "client"):
                        return cand
                # Row below, same column
                if i + 1 < len(rows) and j < len(rows[i + 1]):
                    cand = clean(rows[i + 1][j])
                    if cand:
                        return cand
    return None


def _extract_month_label(rows, sheet_name):
    # 1. Look for MONTH / YEAR label-value pairs in metadata rows
    for i, row in enumerate(rows[:6]):
        for j, v in enumerate(row):
            label = clean(v).lower()
            if label in ("month", "month name", "period", "billing month"):
                month_val = _label_value(rows, i, j)
                if month_val:
                    year_val = _find_year_near(rows, i)
                    if year_val:
                        return f"{month_val.capitalize()} {year_val}"
                    return month_val.capitalize()

    # 2. Look for date objects in first few rows
    for row in rows[:5]:
        for cell in row:
            if is_date(cell):
                return cell.strftime("%B %Y")

    # 3. Regex on sheet name
    m = re.search(r"([A-Z]+)['\s\-]*(\d{2,4})", sheet_name, re.IGNORECASE)
    if m:
        yr = m.group(2)
        if len(yr) == 2:
            yr = "20" + yr
        return f"{m.group(1).capitalize()} {yr}"
    return sheet_name


def _label_value(rows, label_row, label_col):
    """Get the value for a label — check cell below, then cell to the right."""
    # Cell below
    if label_row + 1 < len(rows):
        below = rows[label_row + 1]
        if label_col < len(below) and below[label_col]:
            val = clean(below[label_col])
            if val and val.lower() not in ("month", "year", "period", ""):
                return val
    # Cell to the right
    row = rows[label_row]
    for k in range(label_col + 1, min(label_col + 3, len(row))):
        if row[k]:
            val = clean(row[k])
            if val and val.lower() not in ("year", ""):
                return val
    return None


def _find_year_near(rows, label_row):
    """Find a year value near a MONTH label row."""
    row = rows[label_row]
    for j, v in enumerate(row):
        if clean(v).lower() in ("year", "yr", "calendar year"):
            val = _label_value(rows, label_row, j)
            if val:
                try:
                    yr = int(float(val))
                    if 2000 <= yr <= 2100:
                        return str(yr)
                except (ValueError, TypeError):
                    pass
    # Also check if there's a plain number that looks like a year in the value row
    if label_row + 1 < len(rows):
        for cell in rows[label_row + 1]:
            try:
                yr = int(float(clean(str(cell)))) if cell else 0
                if 2000 <= yr <= 2100:
                    return str(yr)
            except (ValueError, TypeError):
                continue
    return None


# ── Unified data extractor ───────────────────────────────────────────────
def _extract_employees(rows, header_idx, col_map, data_start=None, data_end=None, default_project=None):
    """
    Generic employee extractor — works for ANY layout once you know
    header_idx and col_map {field: col_index}.
    """
    header = rows[header_idx]

    name_idx    = col_map.get("name")
    project_idx = col_map.get("project")
    rate_idx    = col_map.get("billing_rate")
    cost_idx    = col_map.get("cost_rate")
    actual_idx  = col_map.get("actual_hours")
    billable_idx= col_map.get("billable_hours")
    max_idx     = col_map.get("max_hours")
    leaves_idx  = col_map.get("leaves")
    wd_idx      = col_map.get("working_days")

    if name_idx is None:
        name_idx = 0

    # Sub-header scan (for Proofpoint-style fortnights with "Billable"/"Actual" below)
    sub_billable_idx, sub_actual_idx = None, None
    if header_idx + 1 < len(rows):
        sub = rows[header_idx + 1]
        for j, v in enumerate(sub or []):
            lbl = clean(v).lower()
            if "billable" in lbl and sub_billable_idx is None:
                sub_billable_idx = j
            elif "actual" in lbl and sub_actual_idx is None:
                sub_actual_idx = j

    if data_start is None:
        data_start = header_idx + (2 if sub_billable_idx or sub_actual_idx else 1)
    if data_end is None:
        data_end = min(data_start + 50, len(rows) - 1)

    day_cols = _get_daily_date_columns(header)

    employees = {}
    found_any = False
    for r in rows[data_start : data_end + 1]:
        if not r or name_idx >= len(r):
            if found_any:
                break
            continue
        name_val = r[name_idx]
        if not name_val or not clean(name_val):
            if found_any:
                break
            continue
        base_name = clean(name_val)
        name = base_name
        if name.lower() in SKIP_NAMES:
            if found_any:
                break
            continue
        if name in employees:
            suffix = 2
            while f"{base_name} ({suffix})" in employees:
                suffix += 1
            name = f"{base_name} ({suffix})"

        daily = _compute_daily_metrics(r, day_cols) if day_cols else {
            "actual_hours": 0.0,
            "working_days": 0,
            "leave_days": 0,
            "holiday_days": 0,
        }

        row_actual = _safe_float(r, actual_idx) or _safe_float(r, sub_actual_idx)
        actual_hours = daily["actual_hours"] if day_cols else row_actual
        if not actual_hours and row_actual:
            actual_hours = row_actual

        row_working_days = _safe_int(r, wd_idx)
        if _has_cell_value(r, wd_idx):
            working_days = row_working_days
        else:
            working_days = daily["working_days"] if day_cols else row_working_days

        row_leave_days = _safe_int(r, leaves_idx)
        if _has_cell_value(r, leaves_idx):
            leave_days = row_leave_days
        else:
            leave_days = daily["leave_days"] if day_cols else row_leave_days

        # Project
        project = default_project or ""
        if project_idx is not None and project_idx < len(r) and r[project_idx]:
            pv = clean(r[project_idx])
            if pv and not is_date(r[project_idx]):
                project = pv

        employees[name] = {
            "project":       project,
            "billing_rate":  _safe_float(r, rate_idx),
            "cost_rate":     _safe_float(r, cost_idx),
            "actual_hours":  actual_hours,
            "billable_hours":_safe_float(r, billable_idx) or _safe_float(r, sub_billable_idx),
            "expected_hours":_safe_float(r, max_idx),
            "vacation_days": leave_days,
            "leave_days":    leave_days,
            "holiday_days":  daily["holiday_days"],
            "working_days":  working_days,
        }
        found_any = True

    return employees


# ── STRATEGY 1: Pattern-based parsing ────────────────────────────────────
_SKIP_SHEET_NAMES = {"instructions", "template", "readme", "config", "settings",
                     "notes", "data", "lookup", "dropdown", "lists", "master",
                     "reference", "help", "about", "index", "cover", "changelog",
                     "team", "teams", "contacts", "directory", "org", "org chart",
                     "summary", "dashboard", "sheet1", "sheet2", "sheet3"}


def _looks_like_timesheet_sheet(rows, sheet_name):
    """Quick heuristic: does this sheet look like it could contain timesheet data?
    Always trusts actual content (headers) over sheet name."""
    # Check first 30 rows for any header-like content — most reliable signal
    scan = min(30, len(rows))
    for i in range(scan):
        if _is_data_header(rows[i]):
            return True  # has timesheet headers, regardless of sheet name
    # No headers found — use name blacklist to avoid wasting AI on obvious non-data sheets
    if sheet_name.lower().strip() in _SKIP_SHEET_NAMES:
        return False
    # Unknown sheet with no recognizable headers — let AI try
    return True


def _parse_with_patterns(rows):
    """
    Try to parse using pattern matching.
    Handles: fortnights + summary overlay, OR summary-only sheets.
    Returns merged employee dict or {} on failure.
    """
    max_header_scan = min(len(rows), 50)
    all_headers = [
        (i, rows[i]) for i in range(max_header_scan) if _is_data_header(rows[i])
    ]
    if not all_headers:
        return {}

    fortnight_headers = [i for i, h in all_headers if _classify_section(h) == "fortnight"]
    summary_headers   = [i for i, h in all_headers if _classify_section(h) == "summary"]

    # ── A) Fortnight + summary overlay (Proofpoint / TMV style) ──────
    merged = {}
    for h_idx in fortnight_headers:
        header = rows[h_idx]
        col_map = _map_columns(header, ["name", "project", "leaves", "working_days", "max_hours"])
        fort_data = _extract_employees(rows, h_idx, col_map)
        for name, data in fort_data.items():
            if name not in merged:
                merged[name] = {k: 0 for k in ["actual_hours", "billable_hours", "expected_hours", "vacation_days", "leave_days", "holiday_days", "working_days"]}
                merged[name]["project"] = data["project"]
            for k in ["actual_hours", "billable_hours", "expected_hours", "vacation_days", "leave_days", "holiday_days", "working_days"]:
                merged[name][k] += data.get(k, 0)

    # Overlay summary (rates + authoritative totals)
    if summary_headers:
        s_idx = summary_headers[0]
        s_header = rows[s_idx]
        s_col_map = _map_columns(s_header)
        summary_emps = _extract_employees(rows, s_idx, s_col_map,
                                          default_project=_extract_project_from_metadata(rows))
        if merged:
            # Overlay mode: enrich existing fortnight data
            matched_names = 0
            for name in merged:
                sm = _fuzzy_get(name, summary_emps)
                if sm:
                    matched_names += 1
                    for k in ["billing_rate", "cost_rate"]:
                        if sm.get(k):
                            merged[name][k] = sm[k]
                    if sm.get("actual_hours"):
                        merged[name]["actual_hours"] = sm["actual_hours"]
                    if sm.get("billable_hours"):
                        merged[name]["billable_hours"] = sm["billable_hours"]
                    if sm.get("expected_hours"):
                        merged[name]["expected_hours"] = sm["expected_hours"]
                    if sm.get("vacation_days"):
                        merged[name]["vacation_days"] = sm["vacation_days"]
                    if sm.get("leave_days"):
                        merged[name]["leave_days"] = sm["leave_days"]
                else:
                    merged[name].setdefault("billing_rate", 0)
                    merged[name].setdefault("cost_rate", 0)

            # If summary and fortnight sections have zero name overlap,
            # trust summary rows as authoritative for that sheet.
            if matched_names == 0 and summary_emps:
                merged = summary_emps
        else:
            # No fortnights — use summary section as primary data
            merged = summary_emps

    return merged


def _build_mapping_debug(rows):
    """Return lightweight mapping diagnostics for the first detected header row."""
    max_header_scan = min(len(rows), 50)
    for i in range(max_header_scan):
        header = rows[i]
        if not _is_data_header(header):
            continue

        col_map = _map_columns(header)
        fields = {}
        for field, idx in col_map.items():
            if idx is None:
                continue
            label = clean(header[idx]) if idx < len(header) else ""
            fields[field] = {"index": idx, "header": label}

        return {
            "header_row": i + 1,
            "section_type": _classify_section(header),
            "fields": fields,
        }

    return {
        "header_row": None,
        "section_type": "unknown",
        "fields": {},
    }


# ── STRATEGY 2: AI full-sheet analysis ───────────────────────────────────
def _parse_with_ai(rows, sheet_name):
    """
    Ask the LLM to analyze the entire sheet structure and extract employees.
    Returns merged employee dict or {} on failure.
    """
    if not _AI_AVAILABLE or not is_ollama_available():
        return {}

    analysis = ai_analyze_sheet(rows, sheet_name=sheet_name)
    if not analysis:
        return {}

    col_map = analysis["columns"]
    if "name" not in col_map:
        return {}

    default_project = analysis.get("project_name") or _extract_project_from_metadata(rows)
    return _extract_employees(
        rows,
        header_idx=analysis["header_row"],
        col_map=col_map,
        data_start=analysis.get("data_start"),
        data_end=analysis.get("data_end"),
        default_project=default_project,
    )


# ── Name fuzzy-match ──────────────────────────────────────────────────────
def _fuzzy_get(name, lookup):
    if name in lookup:
        return lookup[name]
    name_l = name.lower().strip()
    for key, val in lookup.items():
        key_l = key.lower().strip()
        if name_l in key_l or key_l in name_l:
            return val
    return None


# ── Financial record builder ──────────────────────────────────────────────
def _build_employee_record(name, data, month_label):
    actual_hours   = data.get("actual_hours", 0)
    billable_hours = data.get("billable_hours", 0) or actual_hours
    expected_hours = data.get("expected_hours", 0)
    working_days   = data.get("working_days", 0)
    leave_days     = data.get("leave_days", data.get("vacation_days", 0))
    holiday_days   = data.get("holiday_days", 0)
    vacation_days  = leave_days

    if working_days == 0 and expected_hours > 0:
        working_days = int(expected_hours / HOURS_PER_DAY)

    effective_working_days = max(working_days - vacation_days - holiday_days, 0)
    leave_hours = vacation_days * HOURS_PER_DAY

    billing_rate = data.get("billing_rate", 0)
    cost_rate    = data.get("cost_rate", 0)

    revenue = round(billable_hours * billing_rate, 2) if billing_rate else 0
    cost    = round(actual_hours * cost_rate, 2) if cost_rate else 0
    profit  = round(revenue - cost, 2)
    margin_pct      = round((profit / revenue) * 100, 2) if revenue > 0 else (0 if revenue == 0 and profit == 0 else -100)
    utilisation_pct  = round((actual_hours / expected_hours) * 100, 2) if expected_hours > 0 else 0

    loss_reasons = []
    validation_flags = []

    if profit < 0:
        loss_reasons.append("NEGATIVE_MARGIN")
    if vacation_days >= LEAVE_THRESHOLD:
        loss_reasons.append("HIGH_LEAVE_IMPACT")

    if margin_pct < MARGIN_LOW_THRESHOLD and revenue > 0:
        validation_flags.append("LOW_MARGIN")
    if profit < 0:
        validation_flags.append("NEGATIVE_MARGIN")
    if vacation_days >= LEAVE_THRESHOLD:
        validation_flags.append("HIGH_LEAVE")
    if utilisation_pct < UTILISATION_LOW_THRESHOLD:
        validation_flags.append("LOW_UTILISATION")

    return {
        "employee": name,
        "project": data.get("project", ""),
        "month": month_label,
        "working_days": working_days,
        "leave_days": leave_days,
        "holiday_days": holiday_days,
        "vacation_days": vacation_days,
        "effective_working_days": effective_working_days,
        "actual_hours": actual_hours,
        "billable_hours": billable_hours,
        "expected_hours": expected_hours,
        "leave_hours": leave_hours,
        "billing_rate": billing_rate,
        "cost_rate": cost_rate,
        "revenue": revenue,
        "cost": cost,
        "profit": profit,
        "margin_pct": margin_pct,
        "utilisation_pct": utilisation_pct,
        "loss_reasons": loss_reasons,
        "validation_flags": validation_flags,
        "is_profitable": profit >= 0,
    }


# ── Month filter ──────────────────────────────────────────────────────────
def _month_year_key(label):
    if not label:
        return None
    token_map = {
        "JAN": 1, "JANUARY": 1,
        "FEB": 2, "FEBRUARY": 2,
        "MAR": 3, "MARCH": 3,
        "APR": 4, "APRIL": 4,
        "MAY": 5,
        "JUN": 6, "JUNE": 6,
        "JUL": 7, "JULY": 7,
        "AUG": 8, "AUGUST": 8,
        "SEP": 9, "SEPTEMBER": 9,
        "OCT": 10, "OCTOBER": 10,
        "NOV": 11, "NOVEMBER": 11,
        "DEC": 12, "DECEMBER": 12,
    }
    m = re.search(
        r"(JAN(?:UARY)?|FEB(?:RUARY)?|MAR(?:CH)?|APR(?:IL)?|MAY|JUN(?:E)?|JUL(?:Y)?|AUG(?:UST)?|SEP(?:TEMBER)?|OCT(?:OBER)?|NOV(?:EMBER)?|DEC(?:EMBER)?)\D*(\d{2,4})",
        str(label),
        re.IGNORECASE,
    )
    if not m:
        return None
    month_token = m.group(1).upper()
    year = int(m.group(2))
    if year < 100:
        year += 2000
    month_num = token_map.get(month_token, token_map.get(month_token[:3]))
    if not month_num:
        return None
    return month_num, year


def _match_target_month(sheet_name, target_month):
    if not target_month:
        return True
    target_key = _month_year_key(target_month)
    sheet_key = _month_year_key(sheet_name)
    if target_key and sheet_key:
        return target_key == sheet_key
    def _norm(s):
        s = s.upper().replace("-", "").replace(" ", "").replace("'", "")
        s = re.sub(r"20(\d{2})", r"\1", s)
        return s
    return _norm(target_month) in _norm(sheet_name)


# ── Main entry point ──────────────────────────────────────────────────────
def parse_timesheet(filepath, target_month=None):
    wb = load_workbook(filepath, data_only=True)
    visible_sheetnames = [
        sn for sn in wb.sheetnames if wb[sn].sheet_state == "visible"
    ]

    result = {
        "file": os.path.basename(filepath),
        "sheets": {}
    }

    def _norm_month_text(s):
        s = str(s).upper().replace("-", "").replace(" ", "").replace("'", "")
        s = re.sub(r"20(\d{2})", r"\1", s)
        return s

    exact_month_sheets = set()
    if target_month:
        target_norm = _norm_month_text(target_month)
        exact_month_sheets = {
            sn for sn in visible_sheetnames if _norm_month_text(sn) == target_norm
        }

    # Two-phase sheet selection:
    #   Phase 1: Only sheets matching target_month (fast, precise)
    #   Phase 2: ALL sheets (fallback if Phase 1 found nothing — handles
    #            templates where sheet names don't match the actual month)
    for phase in (1, 2):
        if phase == 2 and result["sheets"]:
            break  # Phase 1 found data, no need for Phase 2

        for sheet_name in visible_sheetnames:
            if sheet_name in result["sheets"]:
                continue  # already processed

            # Phase 1: skip non-matching sheets
            if phase == 1 and target_month:
                if exact_month_sheets:
                    if sheet_name not in exact_month_sheets:
                        continue
                elif not _match_target_month(sheet_name, target_month):
                    continue

            ws = wb[sheet_name]
            rows = [list(r) for r in ws.iter_rows(values_only=True)]

            if len(rows) < 3:
                continue

            # Extract month from sheet CONTENT first (template-agnostic),
            # fall back to sheet name, then to target_month from filename.
            month_label = _extract_month_label(rows, sheet_name)
            if target_month and month_label == sheet_name:
                # Content extraction didn't find anything, use filename hint
                mt = re.match(r"([A-Za-z]+)\D*(\d{2,4})", target_month)
                if mt:
                    yr = mt.group(2)
                    if len(yr) == 2:
                        yr = "20" + yr
                    month_label = f"{mt.group(1).capitalize()} {yr}"

            # Phase 2: skip sheets that don't look like timesheets
            if phase == 2 and not _looks_like_timesheet_sheet(rows, sheet_name):
                continue

            # ── STRATEGY 1: Pattern-based parsing ─────────────────────
            merged = _parse_with_patterns(rows)

            # ── STRATEGY 2: AI fallback ONLY if no headers were found ─
            # If patterns found headers but no employees, AI won't help either
            if not merged and not any(_is_data_header(rows[i]) for i in range(min(30, len(rows)))):
                merged = _parse_with_ai(rows, sheet_name)

            if not merged:
                continue

            mapping_debug = _build_mapping_debug(rows)

            # Quality check: discard if no employee has meaningful data
            has_data = any(
                d.get("actual_hours") or d.get("expected_hours") or d.get("billing_rate")
                for d in merged.values()
            )
            if not has_data:
                continue

            employees = [
                normalize_record(_build_employee_record(name, data, month_label))
                for name, data in merged.items()
            ]

            # ── Per-sheet analytics ───────────────────────────────────
            projects = {}
            for emp in employees:
                proj = emp["project"] or "Unknown"
                if proj not in projects:
                    projects[proj] = {"revenue": 0, "cost": 0, "profit": 0, "employees": 0}
                projects[proj]["revenue"] += emp["revenue"]
                projects[proj]["cost"]    += emp["cost"]
                projects[proj]["profit"]  += emp["profit"]
                projects[proj]["employees"] += 1
            for p in projects.values():
                p["revenue"] = round(p["revenue"], 2)
                p["cost"]    = round(p["cost"], 2)
                p["profit"]  = round(p["profit"], 2)

            total_revenue = round(sum(e["revenue"] for e in employees), 2)
            total_cost    = round(sum(e["cost"] for e in employees), 2)
            total_profit  = round(sum(e["profit"] for e in employees), 2)
            avg_margin    = round((total_profit / total_revenue) * 100, 2) if total_revenue > 0 else 0

            sorted_by_profit = sorted(employees, key=lambda e: e["profit"], reverse=True)
            top_performers = [{"employee": e["employee"], "profit": e["profit"]}
                              for e in sorted_by_profit if e["profit"] > 0][:3]
            low_performers = [{"employee": e["employee"], "profit": e["profit"]}
                              for e in sorted_by_profit if e["profit"] <= 0]

            risks = []
            for e in employees:
                if e["profit"] < 0:
                    risks.append({"employee": e["employee"], "issue": "LOSS_MAKING"})
                if e["vacation_days"] >= LEAVE_THRESHOLD:
                    risks.append({"employee": e["employee"], "issue": "HIGH_LEAVE"})
                if e["utilisation_pct"] < UTILISATION_LOW_THRESHOLD:
                    risks.append({"employee": e["employee"], "issue": "LOW_UTILISATION"})

            result["sheets"][sheet_name] = {
                "template": "generic",
                "mapping_debug": mapping_debug,
                "summary": {
                    "total_employees": len(employees),
                    "total_revenue": total_revenue,
                    "total_cost": total_cost,
                    "total_profit": total_profit,
                    "avg_margin_pct": avg_margin,
                    "total_actual_hours": round(sum(e["actual_hours"] for e in employees), 2),
                    "total_billable_hours": round(sum(e["billable_hours"] for e in employees), 2),
                    "total_working_days": sum(e["working_days"] for e in employees),
                    "total_leave_days": sum(e.get("leave_days", e.get("vacation_days", 0)) for e in employees),
                    "total_holidays": sum(e.get("holiday_days", 0) for e in employees),
                },
                "projects": projects,
                "employees": employees,
                "top_performers": top_performers,
                "low_performers": low_performers,
                "risks": risks,
            }

    total_rev = sum(s["summary"]["total_revenue"] for s in result["sheets"].values())
    total_cst = sum(s["summary"]["total_cost"] for s in result["sheets"].values())
    total_pft = sum(s["summary"]["total_profit"] for s in result["sheets"].values())

    result["overall_summary"] = {
        "total_revenue": round(total_rev, 2),
        "total_cost": round(total_cst, 2),
        "total_profit": round(total_pft, 2),
        "avg_margin_pct": round((total_pft / total_rev) * 100, 2) if total_rev > 0 else 0,
    }

    return result